import openpyxl as xl

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
# cell = sheet["a1"]
# cell1 = sheet.cell(1, 1)
# print(f"total row = {sheet.max_row}")
# print(f"total column = {sheet.max_column}")

for row in range(1, sheet.max_row + 1):
    # print(row.value)
    # for column in range()
    if row == 1:
        for column in range(1, sheet.max_column):
            cell = sheet.cell(row, column + 1)
            if column < sheet.max_column:
                print(f"{cell.value}        ", end="")
            else:
                print(cell.value)
    else:
        for column in range(1, sheet.max_column):
            cell = sheet.cell(row, column + 1)
            if column < sheet.max_column:
                print(f"         {cell.value}", end="")
            else:
                print(cell.value)
    print("")